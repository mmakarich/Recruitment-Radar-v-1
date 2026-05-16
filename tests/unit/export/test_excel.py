"""Testy jednostkowe eksportu Excel."""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from src.export.excel import export_to_excel
from tests.conftest import make_matched_offer


@pytest.fixture
def two_offers(sample_job_offer, sample_salary):
    from datetime import UTC, datetime

    from src.scrapers.base import JobOffer

    offer2 = JobOffer(
        title="Junior React Developer",
        company="Beta Sp. z o.o.",
        portal="nofluffjobs.com",
        url="https://nofluffjobs.com/job/beta-react",
        location="Krakow",
        work_mode="remote",
        seniority="junior",
        tech_stack=("React", "TypeScript"),
        salary=sample_salary,
        published_at=datetime(2026, 5, 12, 10, 0, tzinfo=UTC),
        scraped_at=datetime(2026, 5, 16, 6, 0, tzinfo=UTC),
    )
    return [make_matched_offer(sample_job_offer), make_matched_offer(offer2)]


def test_excel_creates_file(two_offers, tmp_path):
    out = tmp_path / "test.xlsx"
    result = export_to_excel(two_offers, None, out)
    assert result == out
    assert out.exists()
    assert out.stat().st_size > 0


def test_excel_has_three_sheets_with_jd(two_offers, tmp_path, sample_jd):
    out = tmp_path / "test.xlsx"
    export_to_excel(two_offers, sample_jd, out)
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Oferty konkurencji", "Nasza oferta", "Statystyki"}


def test_excel_has_two_sheets_without_jd(two_offers, tmp_path):
    out = tmp_path / "test.xlsx"
    export_to_excel(two_offers, None, out)
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Oferty konkurencji", "Statystyki"}


def test_excel_no_jd_skips_sheet2(two_offers, tmp_path):
    out = tmp_path / "test.xlsx"
    export_to_excel(two_offers, None, out)
    wb = load_workbook(out)
    assert "Nasza oferta" not in wb.sheetnames


def test_excel_hyperlinks_clickable(two_offers, tmp_path):
    out = tmp_path / "test.xlsx"
    export_to_excel(two_offers, None, out)
    wb = load_workbook(out)
    ws = wb["Oferty konkurencji"]
    # Kolumna 14 = Link — sprawdz ze komorki maja hyperlink
    link_cells = [ws.cell(row=r, column=14) for r in range(2, 4)]
    for cell in link_cells:
        assert cell.hyperlink is not None, f"Brak hyperlinka w wierszu {cell.row}"


def test_excel_filters_by_min_score(two_offers, tmp_path):
    out = tmp_path / "test.xlsx"
    # score w make_matched_offer = 75, wiec min_score=80 powinno wykluczyc wszystkie
    export_to_excel(two_offers, None, out, min_score=80)
    wb = load_workbook(out)
    ws = wb["Oferty konkurencji"]
    # Wiersz 2 powinien byc pusty (tylko naglowek)
    assert ws.cell(row=2, column=1).value is None


def test_excel_handles_empty_list(tmp_path):
    out = tmp_path / "empty.xlsx"
    export_to_excel([], None, out)
    wb = load_workbook(out)
    ws = wb["Oferty konkurencji"]
    assert ws.max_row == 1  # tylko naglowek


def test_excel_stats_sheet_has_data(two_offers, tmp_path):
    out = tmp_path / "test.xlsx"
    export_to_excel(two_offers, None, out)
    wb = load_workbook(out)
    ws = wb["Statystyki"]
    # Wiersz 1 = naglowek, wiersz 2+ = portale
    assert ws.cell(row=1, column=1).value == "Portal"
    assert ws.max_row >= 2
