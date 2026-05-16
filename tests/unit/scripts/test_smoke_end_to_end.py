from __future__ import annotations

from pathlib import Path

from docx import Document
from openpyxl import load_workbook

from scripts.smoke_end_to_end import run_smoke


def test_smoke_end_to_end_creates_outputs(tmp_path: Path) -> None:
    summary = run_smoke(tmp_path)

    assert summary["offers_input"] == 3
    assert summary["matched_count"] >= 2
    assert summary["excel_exists"] is True
    assert summary["docx_exists"] is True
    assert summary["top_score"] is not None

    excel_path = Path(summary["excel_path"])
    docx_path = Path(summary["docx_path"])

    assert excel_path.exists()
    assert docx_path.exists()


def test_smoke_excel_is_readable(tmp_path: Path) -> None:
    summary = run_smoke(tmp_path)

    workbook = load_workbook(summary["excel_path"])

    assert "Oferty konkurencji" in workbook.sheetnames
    assert "Nasza oferta" in workbook.sheetnames
    assert "Statystyki" in workbook.sheetnames


def test_smoke_docx_is_readable(tmp_path: Path) -> None:
    summary = run_smoke(tmp_path)

    document = Document(summary["docx_path"])
    full_text = " ".join(paragraph.text for paragraph in document.paragraphs)

    assert "Recruitment Radar" in full_text
    assert "2026-W20" in full_text
