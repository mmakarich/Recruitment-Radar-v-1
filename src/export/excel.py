"""Eksport wyników matchingu do pliku Excel (.xlsx).

Trzy arkusze:
  1. "Oferty konkurencji" - glowna tabela z ofertami i score'ami.
  2. "Nasza oferta"       - sparsowane pola JDParsed (pomijany gdy our_jd=None).
  3. "Statystyki"         - count per portal, median/mean salary.

Uzywamy openpyxl bezposrednio zeby miec pelna kontrole nad formatowaniem:
hyperlinki, conditional formatting, freeze panes, auto-width.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.matching.compare import JDParsed
from src.matching.pipeline import MatchedOffer

_DEEP_NAVY = "1F3A5F"
_WHITE = "FFFFFF"
_LIGHT_GRAY = "F1F5F9"
_BORDER_GRAY = "CBD5E1"

PORTAL_COLORS: dict[str, str] = {
    "justjoin.it": "E8F5E9",
    "nofluffjobs.com": "E3F2FD",
    "rocketjobs.pl": "FFF3E0",
    "theprotocol.it": "F3E5F5",
    "pracuj.pl": "FCE4EC",
}
_DEFAULT_PORTAL_COLOR = "FAFAFA"

_OFFER_COLUMNS: list[tuple[str, int]] = [
    ("Score", 8),
    ("Tytul", 40),
    ("Firma", 25),
    ("Portal (primary)", 18),
    ("Inne portale", 22),
    ("Lokalizacja", 18),
    ("Tryb pracy", 12),
    ("Seniority", 10),
    ("Stawka min", 12),
    ("Stawka max", 12),
    ("Waluta", 8),
    ("Kontrakt", 10),
    ("Okres", 8),
    ("Link", 45),
    ("Opublikowano", 18),
    ("Tech stack", 40),
    ("Duplikaty (portale)", 25),
    ("Pewnosc dedup", 14),
]


def _header_fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def _thin_border() -> Border:
    side = Side(style="thin", color=_BORDER_GRAY)
    return Border(left=side, right=side, top=side, bottom=side)


def _apply_header_row(ws: Worksheet, columns: list[tuple[str, int]]) -> None:
    """Formatuje pierwszy wiersz jako naglowek: bold, navy tlo, bialy tekst, freeze."""
    header_font = Font(bold=True, color=_WHITE, size=10)
    header_fill = _header_fill(_DEEP_NAVY)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (col_name, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = _thin_border()

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def _set_column_widths(ws: Worksheet, columns: list[tuple[str, int]]) -> None:
    for col_idx, (_, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _salary_parts(offer: MatchedOffer) -> tuple[str, str, str, str, str]:
    sal = offer.deduped.primary.salary
    if sal is None:
        return ("—", "—", "—", "—", "—")
    return (str(sal.min), str(sal.max), sal.currency, sal.contract, sal.period)


def _fill_offers_sheet(
    ws: Worksheet,
    matched: list[MatchedOffer],
    min_score: int,
) -> None:
    _apply_header_row(ws, _OFFER_COLUMNS)
    _set_column_widths(ws, _OFFER_COLUMNS)

    filtered = [m for m in matched if (m.match_score.total if m.match_score else 0) >= min_score]

    data_font = Font(size=10)
    link_font = Font(size=10, color="1155CC", underline="single")
    center_align = Alignment(horizontal="center", vertical="top")
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for row_idx, offer in enumerate(filtered, start=2):
        p = offer.deduped.primary
        ms = offer.match_score
        score_val = ms.total if ms is not None else None
        sal_min, sal_max, currency, contract, period = _salary_parts(offer)
        other_portals = ", ".join(port for port in offer.deduped.portals if port != p.portal)

        row_data: list[object] = [
            score_val,
            p.title,
            p.company,
            p.portal,
            other_portals or "—",
            p.location or "—",
            p.work_mode or "—",
            p.seniority or "—",
            sal_min,
            sal_max,
            currency,
            contract,
            period,
            p.url,
            p.published_at.strftime("%Y-%m-%d"),
            ", ".join(p.tech_stack),
            ", ".join(offer.deduped.portals),
            offer.deduped.match_confidence,
        ]

        portal_color = PORTAL_COLORS.get(p.portal, _DEFAULT_PORTAL_COLOR)
        row_fill = _header_fill(portal_color)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _thin_border()
            cell.fill = row_fill
            if col_idx == 14 and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.value = "link"
                cell.font = link_font
                cell.alignment = center_align
            else:
                cell.font = data_font
                if col_idx in (2, 5, 16):
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

        ws.row_dimensions[row_idx].height = 18

    if filtered:
        score_range = f"A2:A{len(filtered) + 1}"
        ws.conditional_formatting.add(
            score_range,
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="FF4444",
                mid_type="num",
                mid_value=50,
                mid_color="FFAA00",
                end_type="num",
                end_value=100,
                end_color="44BB44",
            ),
        )
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_OFFER_COLUMNS))}{len(filtered) + 1}"


def _fill_jd_sheet(ws: Worksheet, our_jd: JDParsed) -> None:
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60

    header_font = Font(bold=True, color=_WHITE, size=11)
    header_fill = _header_fill(_DEEP_NAVY)
    label_font = Font(bold=True, size=10)
    value_font = Font(size=10)

    ws.cell(row=1, column=1, value="Nasza oferta referencyjna").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 22

    sal = our_jd.salary
    fields: list[tuple[str, object]] = [
        ("Tytul", our_jd.title),
        ("Firma", our_jd.company or "—"),
        ("Seniority", our_jd.seniority or "—"),
        ("Lokalizacja", our_jd.location or "—"),
        ("Tryb pracy", our_jd.work_mode or "—"),
        ("Tech stack", ", ".join(our_jd.tech_stack)),
        ("Stawka min", sal.min if sal else "—"),
        ("Stawka max", sal.max if sal else "—"),
        ("Waluta", sal.currency if sal else "—"),
        ("Kontrakt", sal.contract if sal else "—"),
        ("Okres", sal.period if sal else "—"),
    ]

    for row_idx, (label, value) in enumerate(fields, start=2):
        lc = ws.cell(row=row_idx, column=1, value=label)
        lc.font = label_font
        lc.fill = _header_fill(_LIGHT_GRAY)
        lc.border = _thin_border()
        lc.alignment = Alignment(horizontal="right", vertical="center")

        vc = ws.cell(row=row_idx, column=2, value=value)
        vc.font = value_font
        vc.border = _thin_border()
        vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _fill_stats_sheet(ws: Worksheet, matched: list[MatchedOffer]) -> None:
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22

    header_font = Font(bold=True, color=_WHITE, size=10)
    header_fill = _header_fill(_DEEP_NAVY)
    headers = ["Portal", "Liczba ofert", "Sr. score", "Mediana stawki (min PLN/mies)"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center")

    portal_scores: dict[str, list[int]] = {}
    portal_salaries: dict[str, list[int]] = {}

    for offer in matched:
        portal = offer.deduped.primary.portal
        score = offer.match_score.total if offer.match_score else 0
        portal_scores.setdefault(portal, []).append(score)

        salary = offer.deduped.primary.salary
        if salary and salary.currency == "PLN" and salary.period == "month":
            portal_salaries.setdefault(portal, []).append(salary.min)

    for row_idx, (portal, scores) in enumerate(sorted(portal_scores.items()), start=2):
        salaries = portal_salaries.get(portal, [])

        if salaries:
            sorted_salaries = sorted(salaries)
            median_salary: int | str = sorted_salaries[len(sorted_salaries) // 2]
        else:
            median_salary = "—"

        avg_score = round(sum(scores) / len(scores), 1) if scores else 0.0
        row_fill = _header_fill(PORTAL_COLORS.get(portal, _DEFAULT_PORTAL_COLOR))
        data_font = Font(size=10)

        row_values: list[str | int | float] = [portal, len(scores), avg_score, median_salary]

        for col_idx, cell_value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center")

    total_row = len(portal_scores) + 2
    total_values: list[str | int] = ["LACZNIE", len(matched), "", ""]

    for col_idx, total_value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_idx, value=total_value)
        cell.font = Font(bold=True, size=10)
        cell.border = _thin_border()


def export_to_excel(
    matched: list[MatchedOffer],
    our_jd: JDParsed | None,
    output_path: Path,
    min_score: int = 0,
) -> Path:
    """Eksportuje wyniki matchingu do pliku Excel z trzema arkuszami.

    Args:
        matched:     Lista MatchedOffer z pipeline.run_matching().
        our_jd:      Sparsowane nasze ogloszenie (None = pomija Sheet 2).
        output_path: Sciezka do zapisu pliku .xlsx.
        min_score:   Minimalny score do uwzglednienia (0 = wszystkie).

    Returns:
        output_path po zapisaniu pliku.
    """
    wb = Workbook()

    ws_offers = wb.active
    assert ws_offers is not None
    ws_offers.title = "Oferty konkurencji"
    _fill_offers_sheet(ws_offers, matched, min_score)

    if our_jd is not None:
        ws_jd = wb.create_sheet("Nasza oferta")
        _fill_jd_sheet(ws_jd, our_jd)

    ws_stats = wb.create_sheet("Statystyki")
    _fill_stats_sheet(ws_stats, matched)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
